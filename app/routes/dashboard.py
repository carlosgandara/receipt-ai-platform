# app/routes/dashboard.py – Dashboard & Export Blueprint
# Contains /dashboard and /export routes with pagination, filters, charts, JSON export, and Excel export.

import datetime
from collections import defaultdict
import io
import pandas as pd
from openpyxl.utils import get_column_letter   # <-- FIXED: imported for column width

from flask import send_file, Blueprint, request, jsonify, render_template, redirect, url_for, flash

from app.utils.db import (
    find_user_by_email,
    get_user_receipts
)
from app.decorators.auth import token_required
from app.services.image_service import generate_presigned_url
from app.utils.report import generate_report_pdf

dashboard_bp = Blueprint('dashboard', __name__, url_prefix='/')


@dashboard_bp.route('/dashboard')
@token_required
def dashboard():
    """Dashboard with filters, pagination, charts, and table."""
    user = find_user_by_email(request.user_email)
    if not user:
        return redirect(url_for('dashboard'))
    user_id = user.id

    # Get filter parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    category = request.args.get('category')
    merchant_search = request.args.get('merchant')

    # Pagination parameters
    page = int(request.args.get('page', 1))
    limit = int(request.args.get('limit', 25))
    if page < 1:
        page = 1
    if limit < 1:
        limit = 25
    if limit > 100:
        limit = 100

    # Fetch all filtered receipts (we need total count and stats)
    all_filtered = get_user_receipts(
        user_id,
        start_date=start_date,
        end_date=end_date,
        category=category,
        merchant=merchant_search,
        include_deleted=False
    )

    # Summary stats
    total_receipts = len(all_filtered)
    total_spent = sum(r.total for r in all_filtered) if all_filtered else 0
    avg_spent = total_spent / total_receipts if total_receipts else 0
    max_receipt = max(all_filtered, key=lambda x: x.total) if all_filtered else None
    min_receipt = min(all_filtered, key=lambda x: x.total) if all_filtered else None

    # Paginate
    offset = (page - 1) * limit
    paginated = all_filtered[offset:offset + limit]

    # Build records list with presigned URLs for S3 images
    records_list = []
    for r in paginated:
        image_url = None
        if r.s3_key:
            image_url = generate_presigned_url(r.s3_key)
        else:
            image_url = r.image_path
        records_list.append({
            'id': r.id,
            'merchant': r.merchant,
            'date': r.date,
            'time': r.time,
            'total': r.total,
            'category': r.category,
            'comment': r.comment,
            'image_path': image_url,
            'created_at': r.created_at
        })

    total_pages = (total_receipts + limit - 1) // limit if total_receipts > 0 else 1

    # Chart data (from all filtered, not paginated)
    cat_totals = defaultdict(float)
    for r in all_filtered:
        cat = r.category or 'OTHER'
        cat_totals[cat] += r.total

    weekly = defaultdict(float)
    for r in all_filtered:
        if r.date:
            try:
                dt = datetime.datetime.strptime(r.date, '%Y-%m-%d')
                week_start = dt - datetime.timedelta(days=dt.weekday())
                key = week_start.strftime('%Y-%m-%d')
                weekly[key] += r.total
            except:
                pass
    sorted_weekly = sorted(weekly.items())
    dates = [item[0] for item in sorted_weekly]
    weekly_totals = [item[1] for item in sorted_weekly]

    chart_data = {
        'categories': list(cat_totals.keys()),
        'cat_values': [cat_totals[c] for c in cat_totals],
        'dates': dates,
        'weekly_totals': weekly_totals
    }

    # Get unique merchants for filter dropdown
    all_user_receipts = get_user_receipts(user_id, include_deleted=False)
    merchants = sorted(set(r.merchant for r in all_user_receipts if r.merchant))

    # Template data
    template_data = {
        'records': records_list,
        'chart_data_json': chart_data,
        'total_receipts': total_receipts,
        'total_spent': total_spent,
        'avg_spent': avg_spent,
        'max_receipt': max_receipt,
        'min_receipt': min_receipt,
        'merchants': merchants,
        'selected_category': category or 'ALL',
        'selected_merchant': merchant_search or '',
        'start_date': start_date or '',
        'end_date': end_date or '',
        'user_email': request.user_email,
        'page': page,
        'limit': limit,
        'total_pages': total_pages
    }

    # If AJAX request, return JSON
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        json_records = []
        for r in paginated:
            image_url = None
            if r.s3_key:
                image_url = generate_presigned_url(r.s3_key)
            else:
                image_url = r.image_path
            json_records.append({
                'id': r.id,
                'merchant': r.merchant,
                'date': r.date,
                'time': r.time,
                'total': r.total,
                'category': r.category,
                'comment': r.comment,
                'image_path': image_url,
                'created_at': r.created_at
            })
        return jsonify({
            'records': json_records,
            'total_receipts': total_receipts,
            'total_spent': total_spent,
            'avg_spent': avg_spent,
            'max_receipt': max_receipt,
            'min_receipt': min_receipt,
            'page': page,
            'limit': limit,
            'total_pages': total_pages
        })

    return render_template('dashboard.html', **template_data)


@dashboard_bp.route('/export')
@token_required
def export_json():
    """Export filtered data as JSON (same filters as dashboard)."""
    user = find_user_by_email(request.user_email)
    if not user:
        return jsonify({"error": "User not found"}), 401
    user_id = user.id

    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    category = request.args.get('category')
    merchant_search = request.args.get('merchant')

    records = get_user_receipts(
        user_id,
        start_date=start_date,
        end_date=end_date,
        category=category,
        merchant=merchant_search,
        include_deleted=False
    )

    data = [{
        'id': r.id,
        'merchant': r.merchant,
        'date': r.date,
        'time': r.time,
        'subtotal': r.subtotal,
        'tax': r.tax,
        'total': r.total,
        'payment_method': r.payment_method,
        'category': r.category,
        'comment': r.comment,
        'created_at': r.created_at.isoformat() if r.created_at else None
    } for r in records]

    return jsonify(data)


@dashboard_bp.route('/report')
@token_required
def download_report():
    """Download a PDF report for the given month/year."""
    user = find_user_by_email(request.user_email)
    if not user:
        return jsonify({"error": "User not found"}), 401

    month = request.args.get('month')
    year = request.args.get('year')
    if not month or not year:
        # fallback to current month
        now = datetime.datetime.now()
        month = now.strftime('%Y-%m')
        year = str(now.year)

    pdf_data = generate_report_pdf(user.id, month, year, user.email)
    return send_file(
        io.BytesIO(pdf_data),
        as_attachment=True,
        download_name=f"report-{month}.pdf",
        mimetype='application/pdf'
    )


# ============================================================
# FIXED Excel Export – with proper column letters & organization
# ============================================================
@dashboard_bp.route('/export/excel')
@token_required
def export_excel():
    """Export filtered data as Excel (.xlsx) with organized summaries."""
    user = find_user_by_email(request.user_email)
    if not user:
        return jsonify({"error": "User not found"}), 401
    user_id = user.id

    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    category = request.args.get('category')
    merchant_search = request.args.get('merchant')

    records = get_user_receipts(
        user_id,
        start_date=start_date,
        end_date=end_date,
        category=category,
        merchant=merchant_search,
        include_deleted=False
    )

    # ----- 1. Build Receipts sheet data (sorted by date descending) -----
    data = []
    for r in records:
        data.append({
            'Date': r.date or '',
            'Merchant': r.merchant or '',
            'Total': float(r.total) if r.total else 0,
            'Category': r.category or '',
            'Deduction Category': getattr(r, 'deduction_category', '') or '',
            'Comment': r.comment or '',
            'Payment Method': r.payment_method or '',
            'Subtotal': float(r.subtotal) if r.subtotal else 0,
            'Tax': float(r.tax) if r.tax else 0,
        })

    df_receipts = pd.DataFrame(data)
    # Sort by date descending (newest first) – convert date to datetime for proper sorting
    df_receipts['Date_parsed'] = pd.to_datetime(df_receipts['Date'], errors='coerce')
    df_receipts = df_receipts.sort_values('Date_parsed', ascending=False).drop('Date_parsed', axis=1)

    # ----- 2. Category Summary (sorted descending) -----
    cat_summary = df_receipts.groupby('Category', as_index=False)['Total'].sum()
    cat_summary.columns = ['Category', 'Total Spent']
    cat_summary = cat_summary[cat_summary['Category'] != '']  # remove empty
    cat_summary = cat_summary.sort_values('Total Spent', ascending=False)

    # Add a Grand Total row
    grand_total_cat = cat_summary['Total Spent'].sum()
    cat_summary.loc[len(cat_summary)] = ['GRAND TOTAL', grand_total_cat]

    # ----- 3. Deduction Summary (sorted descending) -----
    ded_summary = df_receipts.groupby('Deduction Category', as_index=False)['Total'].sum()
    ded_summary.columns = ['Deduction Category', 'Total Spent']
    ded_summary = ded_summary[ded_summary['Deduction Category'] != '']  # remove empty
    ded_summary = ded_summary.sort_values('Total Spent', ascending=False)

    # Add a Grand Total row
    grand_total_ded = ded_summary['Total Spent'].sum()
    ded_summary.loc[len(ded_summary)] = ['GRAND TOTAL', grand_total_ded]

    # ----- 4. Write Excel with formatting -----
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write all sheets
        df_receipts.to_excel(writer, index=False, sheet_name='Receipts')
        cat_summary.to_excel(writer, index=False, sheet_name='Category Summary')
        if not ded_summary.empty:
            ded_summary.to_excel(writer, index=False, sheet_name='Deduction Summary')

        # Get workbook and adjust each sheet
        workbook = writer.book
        sheet_mapping = {
            'Receipts': df_receipts,
            'Category Summary': cat_summary,
            'Deduction Summary': ded_summary
        }

        for sheet_name, df in sheet_mapping.items():
            if sheet_name not in writer.sheets:
                continue
            worksheet = writer.sheets[sheet_name]

            # Set column widths using openpyxl's get_column_letter
            for i, col in enumerate(df.columns, 1):
                max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
                col_letter = get_column_letter(i)
                worksheet.column_dimensions[col_letter].width = min(max_len, 50)

            # Bold headers
            for cell in worksheet[1]:
                cell.font = cell.font.copy(bold=True)

            # Format currency for the second column in summary sheets
            if sheet_name != 'Receipts':
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=2)
                    cell.number_format = '$#,##0.00'

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"receipts_export_{datetime.datetime.now().strftime('%Y%m%d')}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )